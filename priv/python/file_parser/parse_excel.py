"""Parse Excel (.xlsx) and CSV files into structured dicts.

Returns:
    {"type": "structured", "sheets": [{"name": str, "columns": [str], "rows": [dict], "row_count": int}]}
    {"type": "error", "message": str}
"""

import csv
import io
import os
import sys


def _parse_csv(file_path):
    """Parse a CSV file with encoding detection and delimiter sniffing."""
    try:
        import chardet
    except ImportError:
        chardet = None

    # Read raw bytes for encoding detection
    with open(file_path, "rb") as f:
        raw = f.read()

    # Detect encoding
    if chardet is not None:
        detected = chardet.detect(raw)
        encoding = detected.get("encoding") or "utf-8"
    else:
        encoding = "utf-8"

    text = raw.decode(encoding, errors="replace")

    # Sniff delimiter
    try:
        sample = text[:8192]
        dialect = csv.Sniffer().sniff(sample)
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)

    if not all_rows:
        return {"type": "structured", "sheets": [{"name": "Sheet1", "columns": [], "rows": [], "row_count": 0}]}

    headers = all_rows[0]
    rows = []
    for row in all_rows[1:]:
        # Skip empty rows
        if not any(cell.strip() for cell in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            row_dict[header] = row[i] if i < len(row) else ""
        rows.append(row_dict)

    return {
        "type": "structured",
        "sheets": [
            {
                "name": os.path.splitext(os.path.basename(file_path))[0],
                "columns": headers,
                "rows": rows,
                "row_count": len(rows),
            }
        ],
    }


def _parse_xlsx(file_path):
    """Parse an Excel .xlsx file, reading ALL sheets."""
    try:
        import openpyxl
    except ImportError:
        return {"type": "error", "message": "openpyxl is not installed. Run: pip install openpyxl"}

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = []
            for row in ws.iter_rows(values_only=True):
                raw_rows.append(list(row))

            if not raw_rows:
                sheets.append({"name": sheet_name, "columns": [], "rows": [], "row_count": 0})
                continue

            # First row is headers
            headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(raw_rows[0])]
            rows = []
            for raw_row in raw_rows[1:]:
                # Skip empty rows
                if not any(cell is not None for cell in raw_row):
                    continue
                row_dict = {}
                for i, header in enumerate(headers):
                    val = raw_row[i] if i < len(raw_row) else None
                    row_dict[header] = val
                rows.append(row_dict)

            sheets.append({
                "name": sheet_name,
                "columns": headers,
                "rows": rows,
                "row_count": len(rows),
            })
    finally:
        wb.close()

    return {"type": "structured", "sheets": sheets}


def parse(file_path, mime_type=""):
    """Parse an Excel or CSV file.

    Args:
        file_path: Path to the file.
        mime_type: Optional MIME type hint. If empty, inferred from extension.

    Returns:
        A dict with type "structured" or "error".
    """
    try:
        ext = os.path.splitext(file_path)[1].lower()

        if mime_type in ("text/csv", "application/csv") or ext == ".csv":
            return _parse_csv(file_path)
        elif mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ) or ext in (".xlsx", ".xls"):
            return _parse_xlsx(file_path)
        else:
            # Fallback: try CSV if text-like, else try xlsx
            try:
                return _parse_xlsx(file_path)
            except Exception:
                return _parse_csv(file_path)
    except Exception as e:
        return {"type": "error", "message": str(e)}


if __name__ == "__main__":
    import json

    if len(sys.argv) < 2:
        print("Usage: python parse_excel.py <file_path> [mime_type]")
        sys.exit(1)

    path = sys.argv[1]
    mime = sys.argv[2] if len(sys.argv) > 2 else ""
    result = parse(path, mime)
    print(json.dumps(result, indent=2, default=str))
