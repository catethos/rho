#!/usr/bin/env python3
"""
Extract FSF (Future Skills Framework) Excel file into framework rows.

Input: FSF-Job-Roles-and-Skills_Master-Database.xlsx
Output: JSON array of rows matching FrameworkRow schema

Schema: role, category, cluster, skill_name, skill_description,
        level (1-5), level_name, level_description, skill_code
"""

import openpyxl
import json
import sys

def extract_fsf(file_path, output_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # --- Step 1: Extract skills + proficiency levels from "Skills to Job Roles Mapping" ---
    ws = wb["Skills to Job Roles Mapping"]

    # Find header rows (row 6 = department groups, row 7 = role names + PL headers)
    # Columns: A=No, B=Category, C=Cluster, D=Skill Name, E=Definition, F-J=PL1-PL5, K=blank, L+=roles

    # Identify PL columns (row 7 has "PL 1", "PL 2", etc.)
    pl_columns = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=7, column=col).value
        if val and "PL" in str(val).strip():
            # Extract level number
            level_str = str(val).strip().replace("PL", "").strip()
            try:
                level = int(level_str)
                pl_columns[level] = col
            except ValueError:
                pass

    print(f"Found PL columns: {pl_columns}", file=sys.stderr)

    # Identify role columns (row 7, after PL columns)
    role_columns = []  # [(col_index, role_name)]
    for col in range(max(pl_columns.values()) + 1, ws.max_column + 1):
        role_name = ws.cell(row=7, column=col).value
        if role_name and str(role_name).strip():
            role_columns.append((col, str(role_name).strip()))

    print(f"Found {len(role_columns)} role columns", file=sys.stderr)

    # Extract skills (rows 8+)
    skills = []
    for row in range(8, ws.max_row + 1):
        skill_name = ws.cell(row=row, column=4).value
        if not skill_name or not str(skill_name).strip():
            continue

        skill = {
            "skill_code": str(ws.cell(row=row, column=1).value or "").strip(),
            "category": str(ws.cell(row=row, column=2).value or "").strip(),
            "cluster": str(ws.cell(row=row, column=3).value or "").strip(),
            "skill_name": str(skill_name).strip(),
            "skill_description": str(ws.cell(row=row, column=5).value or "").strip(),
            "proficiency_levels": {},
            "mapped_roles": [],
        }

        # Extract proficiency level descriptions
        for level, col in pl_columns.items():
            desc = ws.cell(row=row, column=col).value
            if desc and str(desc).strip():
                skill["proficiency_levels"][level] = str(desc).strip()

        # Extract role mappings (Y = this role requires this skill)
        for col, role_name in role_columns:
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().upper() == "Y":
                skill["mapped_roles"].append(role_name)

        skills.append(skill)

    print(f"Extracted {len(skills)} skills", file=sys.stderr)

    # --- Step 2: Build framework rows ---
    # For each skill:
    #   - Base rows (no role): 1 row per proficiency level
    #   - Role rows: for each mapped role, 1 row per proficiency level

    rows = []

    for skill in skills:
        pl = skill["proficiency_levels"]
        base = {
            "skill_code": skill["skill_code"],
            "category": skill["category"],
            "cluster": skill["cluster"],
            "skill_name": skill["skill_name"],
            "skill_description": skill["skill_description"],
        }

        # Base rows (no role) — the skill definition with all proficiency levels
        for level in sorted(pl.keys()):
            rows.append({
                **base,
                "role": "",
                "level": level,
                "level_name": f"PL {level}",
                "level_description": pl[level],
            })

        # Role-specific rows — for each role that has Y, replicate all proficiency levels
        for role_name in skill["mapped_roles"]:
            for level in sorted(pl.keys()):
                rows.append({
                    **base,
                    "role": role_name,
                    "level": level,
                    "level_name": f"PL {level}",
                    "level_description": pl[level],
                })

    print(f"Generated {len(rows)} framework rows", file=sys.stderr)
    print(f"  Base rows (no role): {sum(1 for r in rows if r['role'] == '')}", file=sys.stderr)
    print(f"  Role rows: {sum(1 for r in rows if r['role'] != '')}", file=sys.stderr)
    print(f"  Unique roles: {len(set(r['role'] for r in rows if r['role'] != ''))}", file=sys.stderr)

    # --- Step 3: Write output ---
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False)

    print(f"Written to {output_path}", file=sys.stderr)
    return len(rows)


if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else "FSF-Job-Roles-and-Skills_Master-Database.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "/tmp/fsf_framework_rows.json"
    extract_fsf(file_path, output_path)
